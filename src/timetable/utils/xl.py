from django.conf import settings
import openpyxl

class Xl:
    __COLUMNS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    def __init__(self, filename: str):
        self.filename = filename
        self.filepath = str(settings.TABLES_DIR / filename) + '.xlsx'

    def create_workbook(self):
        openpyxl.Workbook().save(self.filepath)
    
    def load_workbook(self):
        """Open workbook and get access to worksheet."""
        try:
            workbook = openpyxl.load_workbook(self.filepath)
            worksheet = workbook.active
            return workbook, worksheet
        except Exception as e:
            print('Failed', e)
            exit()

    def setup(self, timeslots, locations = None, rooms = None):
        """Setup table by writing titles for rows (locations<rooms>) and columns (timeslots)."""
        self.create_workbook()
        workbook, worksheet = self.load_workbook()
        if locations and rooms or not locations and not rooms:
            raise Exception('Choose setup by locations or rooms only!')
        
        # Setup columns.
        col = 1
        if locations:
            for location in locations:
                for room in location.room_set.all():
                    worksheet[self.__COLUMNS[col] + '1'] = location.code + '-' + room.code
                    col += 1
        if rooms:
            for room in rooms:
                worksheet[self.__COLUMNS[col] + '1'] = room.location.code + '-' + room.code
                col += 1
        
        # Setup rows.
        row = 2
        counter = 1
        for timeslot in timeslots:
            worksheet['A' + str(row)] = 'T' + str(counter) + '-' + timeslot.code
            row += 1
            counter += 1
        workbook.save(self.filepath)
    
    def write(self, timetable):
        """Write timetable to xlsx file."""
        workbook, worksheet = self.load_workbook() 
        for key, value in timetable.items():
            room, timeslot = value.values()
            row = self.find_in_row(tuple(worksheet.rows), timeslot.code)
            col = self.find_in_column(tuple(worksheet.columns), room.code)
            cell = col + row
            worksheet[cell] = key.course.code + '-' + str(key.number)
        workbook.save(self.filepath)
    
    def find_in_row(self, rows, value):
        """Find value in rows."""
        for n, row in enumerate(rows):
            for cell in row:
                if cell.value == None: continue
                if cell.value == value or value in cell.value:
                    return str(n + 1)

    def find_in_column(self, columns, value):
        for n, column in enumerate(columns):
            for cell in column:
                if cell.value == None: continue
                if cell.value == value or value in cell.value:
                    return self.__COLUMNS[n]
